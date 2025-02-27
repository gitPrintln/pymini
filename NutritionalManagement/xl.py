import pandas as pd
import numpy as np
from openpyxl import Workbook


# 1. xlsx 파일 경로
file_path = 'NutritionalManagement\\NMdata\\2020 한국인 영양소 섭취 기준표.xlsx'

# 2. Excel 파일 읽기 (첫 번째 시트를 기본으로 읽음)
# openpyxl 라이브러리 사용
# header=[0,1] 옵션을 사용하면 계층적 컬럼 구조
# 여러 줄의 헤더를 가질 경우
# 각 컬럼이 (대분류, 소분류) 형태로 계층 구조를 가짐
df = pd.read_excel(file_path, header=[0, 1], engine='openpyxl')

# 두 줄의 헤더를 연결하여 한 줄의 헤더로 만들기
df.columns = ['_'.join(col).strip() for col in df.columns]

# 컬럼명 \n 제거
df.columns = df.columns.str.replace("\n", "")
# 컬럼명에서 .1, .2, .3 패턴이 있는 컬럼 찾기
#cols_to_remove = [col for col in df.columns if ".1" in col or ".2" in col or ".3" or ".4" in col]
# 해당 컬럼 제거
#df = df.drop(columns=cols_to_remove)

 # \n으로 나누어져 있는 열들을 개별 행으로 분리하는 함수
 # 한 행에 한 줄씩 들어가도록 뒤의 값들도 마찬가지로 해줌
 # 중간 중간에 성별, 연령, 지방,비타민 이런식으로 header부분이 들어가므로 이부분에서 잠깐
 # 끊었다가 다시 이어서 붙이려고 체크포인트 지점 만들어줌
check_point_idx = 0
def split_and_expand(df, check_point_idx):
    temp_expanded_df = pd.DataFrame()
    for idx, row in df.iloc[check_point_idx:].iterrows():
        # 각 행에 대해 '\n'으로 나누고 빈 셀은 제거
        expanded_rows = []
        for cell in row:
            if isinstance(cell, str) and '\n' in cell:
                expanded_rows.append(cell.split('\n'))
            else:
                expanded_rows.append([cell])
        # 최대 길이에 맞춰 열을 확장
        max_len = max(len(items) for items in expanded_rows)
        expanded_rows = [items + [np.nan] * (max_len - len(items)) for items in expanded_rows]
        # 새로운 DataFrame으로 변환
        temp_df = pd.DataFrame(expanded_rows).T
        temp_df.columns = df.columns
        temp_expanded_df = pd.concat([temp_expanded_df, temp_df], ignore_index=True)
        # 값이 성별인 경우 끊었다가 다시 해야함.
        if '성별' in row.values:
            check_point_idx = idx + 1 # 한 단계 성별 라인의 행은 건너서 넘어가고
            return check_point_idx, temp_expanded_df
    
    return check_point_idx, temp_expanded_df

expanded_df = pd.DataFrame()  # 임시 DataFrame 생성
for idx, row in df.iterrows(): # 0 ~ 끝 인덱스까지
    if idx == check_point_idx:
        # DataFrame 확장
        check_point_idx, piece_expanded_df = split_and_expand(df, check_point_idx)
        expanded_df = pd.concat([expanded_df, piece_expanded_df], ignore_index=True)  # expanded_df에 덧붙이기
    else: # 현재 for문에서도 인덱스를 맞춰줌. 체크포인트부터 다시 시작하기 위해서
        idx += 1

# 첫 성별 부분, 남자면 쭉 남자, 여자면 쭉 여자, 유아면 쭉 유아 빈부분 채워넣기
#for col in expanded_df.columns:
expanded_df["성별_Unnamed: 0_level_1"] = expanded_df["성별_Unnamed: 0_level_1"].fillna(method='ffill')

expanded_df.reset_index(drop=True, inplace=True)

# 수유부 아래에 3 행을 처리(성별, 연령, 셀레늄_평균필요량) 이런 식으로 처리
target_cell = "수유부"
# target_cell과 정확히 일치하는 셀 찾기
matches = expanded_df.eq(target_cell)  # target_cell과 같은 셀이면 True, 아니면 False
# 빠르게 위치만 찾기(수유부 값 아래 두 번째 행에 값이 있다면 작업해주기)
matched_cells = matches.stack().loc[lambda x: x].index.tolist()
target_idx = [index + 2 for index, label in matched_cells]
for idx in target_idx[:-1]:
    temp_header = None
    for x, data in expanded_df.iloc[idx, 2:].items():
        if pd.notna(data):
            if(pd.notna(expanded_df.loc[idx-1, x])):
                temp_header = expanded_df.loc[idx-1, x]
                expanded_df.loc[idx-1, x] = str(expanded_df.loc[idx-1, x]) + "_" + str(expanded_df.loc[idx, x]) + str(expanded_df.loc[idx+1, x])
            else:
                expanded_df.loc[idx-1, x] = temp_header + "_" + str(expanded_df.loc[idx, x]) + str(expanded_df.loc[idx+1, x])
    
delete_idxs = []        
for idx in target_idx:
    delete_idxs.append(idx)
    if idx + 1 < len(expanded_df):
        delete_idxs.append(idx + 1)
delete_idxs.remove(461)
delete_idxs.append(246)
expanded_df.drop(delete_idxs, inplace=True)

# 스프레드 시트에 테이블별로 나누기
def split_and_save(df):
    cnt = 0
    current_data = []
    wb = Workbook() # 새로운 Excel 파일
    ws = wb.active # 현재 활성화된 시트를 가져옴
    ws.title = f"Sheet{cnt + 1}" # 현재 활성화된 시트의 이름 변경
    
    for idx, row in expanded_df.iterrows():
        # 첫 컬럼에서 "성별"이 나타날 때마다 데이터를 새로운 시트로 나누기
        if row["성별_Unnamed: 0_level_1"] == "성별":
            # 데이터가 있으면 현재 시트에 추가
            if current_data:
                for row_data in current_data:
                    ws.append(row_data)
            
            # 새로운 시트 생성
            cnt += 1
            ws = wb.create_sheet(f"Sheet{cnt + 1}") if cnt > 0 else wb.active
            current_data = [row.tolist()]  # 현재 데이터 새로운 데이터로 설정
        
        # 성별이 아닌경우 데이터 추가
        else: 
            current_data.append(row.tolist())
    
    # 마지막 데이터 저장
    if current_data:
        for row_data in current_data:
            ws.append(row_data)
    
    # 빈 컬럼 삭제
    import math
    sheet_names = wb.sheetnames
    for sheet in sheet_names:
        current_sheet = wb[sheet] # 해당 시트 가져옴
        max_column = current_sheet.max_column # 컬럼 길이
        for col in range(max_column, 0, -1): # 1씩 감소하면서 오른쪽 부터 열을 삭제하면서 idx 바뀌는 거 방지
            # 열의 모든 셀값 가져와서
            column_cells = [current_sheet.cell(row=row, column=col).value for row in range(1, current_sheet.max_row + 1)]
            # 모두 비어있는지 확인 
            # column_cells은 값을 확인해보면 NaN은 주로 pandas나 numpy에서 수치 계산을 할 때 발생하는 값으로 빈값이 nan이 발생
            # 하지만 아래 조건 검사에서는 openpyxl의 엑셀파일로 읽을 때는 None가 발생 
            # 즉 nan이 있는 값에서 None를 확인은 잘 작동이 안되는 상황이어서 nan을 제대로 확인하려면
            # math.isnan을 사용하여 nan은 float형식이기 때문에 이걸로 처리해줘야함.
            # 만약 pandas에서 Nan처리를 할거라면 dropna()라는 간단한 메서드가 있기는함.
            # 여기서는 openpyxl을 사용했기 때문에 아래와 같이 조건 검사
            if all(cell is None or (isinstance(cell, float) and math.isnan(cell)) for cell in column_cells):
                current_sheet.delete_cols(col) # 삭제
                
    # 첫 번째 시트 날라간 header 추가
    first_sheet = wb.worksheets[0]
    first_sheet_header = ['성별', '연령', '에너지(kcal/일)_필요추정량', '탄수화물(g/일)_평균필요량', 
                        '탄수화물(g/일)_권장섭취량', '탄수화물(g/일)_충분섭취량', '식이섬유(g/일)_충분섭취량']
    first_sheet.insert_rows(1)
    for col, header in enumerate(first_sheet_header, start=1):
        first_sheet.cell(row=1, column=col, value=header)
    
    # 워크북 최종 저장
    wb.save("final_xl.xlsx")

# 시트 별로 하나에 한 테이블씩 들어가도록 나누기/빈 컬럼 삭제/사라진 첫 컬럼 header 생성
split_and_save(expanded_df)



# 최종 엑셀 쓰기
# expanded_df.to_excel('modified_file8.xlsx', index=False)